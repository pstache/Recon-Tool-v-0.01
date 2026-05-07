"""
End-of-month Excel reporting.

Per spec + clarifications #18 / #19:
  * "End-of-month" = calendar month based on settlement_date (#19).
  * Top-5 deviations: per ISIN, summed across institutions (#18).
  * Deviation metric = absolute quantity difference between paired internal
    and external rows. Where one side is missing (e.g. partial without
    counterpart), deviation = absolute value of the present row's quantity.

Sheets generated:
  Summary        — high-level counts (matched/unmatched per source)
  Top Deviations — top 5 (ISIN, deviation_qty) — one block per partial reason
  Auto Matches   — every auto-matched pair in the month
  Partial Matches — every partial match in the month
  Unmatched      — every still-unmatched active transaction whose settlement
                   falls in the month
"""

from __future__ import annotations

import datetime as dt
from decimal import Decimal
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from storage.repository import Repository


HEADER_FILL = PatternFill("solid", fgColor="2E75B6")
HEADER_FONT = Font(bold=True, color="FFFFFF", name="Arial")
DATA_FONT   = Font(name="Arial")


def generate_monthly_report(
    repo: Repository,
    *,
    year: int,
    month: int,
    out_path: str | Path,
) -> Path:
    """Generate a monthly reconciliation report. Returns the output path."""
    out_path = Path(out_path)
    out_path.parent.mkdir(parents=True, exist_ok=True)

    period_start = dt.date(year, month, 1)
    period_end   = dt.date(year + (1 if month == 12 else 0),
                           1 if month == 12 else month + 1, 1) - dt.timedelta(days=1)

    wb = Workbook()
    # Default sheet -> Summary
    ws = wb.active
    ws.title = "Summary"
    _write_summary(ws, repo, period_start, period_end)

    _write_top_deviations(wb.create_sheet("Top Deviations"),
                          repo, period_start, period_end)
    _write_pairs_sheet(wb.create_sheet("Auto Matches"),
                       repo, period_start, period_end, status="auto_matched")
    _write_pairs_sheet(wb.create_sheet("Partial Matches"),
                       repo, period_start, period_end, status="partial")
    _write_unmatched_sheet(wb.create_sheet("Unmatched"),
                           repo, period_start, period_end)

    wb.save(out_path)
    return out_path


# ---------------------------------------------------------------- Summary

def _write_summary(ws, repo: Repository, start: dt.date, end: dt.date) -> None:
    ws["A1"] = f"Reconciliation Report — {start.strftime('%B %Y')}"
    ws["A1"].font = Font(name="Arial", size=14, bold=True)
    ws["A2"] = f"Settlement-date window: {start.isoformat()} … {end.isoformat()}"
    ws["A2"].font = Font(name="Arial", italic=True)

    rs = repo.db.execute("""
        SELECT
            COALESCE(m.status, 'unmatched') AS status,
            t.source,
            COUNT(*) AS n
        FROM transactions_all t
        LEFT JOIN matches m ON
                t.transaction_id = m.internal_id
             OR t.transaction_id = m.external_id
        WHERE t.is_active = TRUE
          AND t.settlement_date BETWEEN ? AND ?
        GROUP BY 1, 2
        ORDER BY 1, 2
    """, [start, end]).fetchall()

    headers = ["Status", "Source", "Count"]
    _write_header(ws, row=4, headers=headers)
    for i, (status, source, n) in enumerate(rs, start=5):
        ws.cell(row=i, column=1, value=status).font = DATA_FONT
        ws.cell(row=i, column=2, value=source).font = DATA_FONT
        ws.cell(row=i, column=3, value=n).font = DATA_FONT

    _autosize(ws, max_col=3)


# -------------------------------------------------------- Top Deviations

def _write_top_deviations(ws, repo: Repository, start: dt.date, end: dt.date) -> None:
    """Top 5 ISINs by absolute quantity deviation between paired rows.

    For each pair (matched or partial) where both sides exist, deviation =
    |qty_internal - qty_external|. We sum across institutions per ISIN
    (clarification #18) and take the top 5 overall.

    For partial rows where external_id is null (e.g. transfer split flags),
    the deviation falls back to |qty_internal|.
    """
    ws["A1"] = "Top 5 ISIN deviations (per ISIN, summed across institutions)"
    ws["A1"].font = Font(name="Arial", size=14, bold=True)
    ws["A2"] = f"Settlement window: {start.isoformat()} … {end.isoformat()}"
    ws["A2"].font = Font(name="Arial", italic=True)

    rs = repo.db.execute("""
        WITH pairs AS (
            SELECT
                i.isin,
                i.quantity AS i_qty,
                COALESCE(e.quantity, 0) AS e_qty
            FROM matches m
            JOIN transactions_all i ON m.internal_id = i.transaction_id
            LEFT JOIN transactions_all e ON m.external_id = e.transaction_id
            WHERE i.settlement_date BETWEEN ? AND ?
        )
        SELECT
            isin,
            SUM(ABS(i_qty - e_qty)) AS total_deviation_qty,
            COUNT(*) AS n_pairs
        FROM pairs
        GROUP BY isin
        ORDER BY total_deviation_qty DESC
        LIMIT 5
    """, [start, end]).fetchall()

    headers = ["ISIN", "Total deviation (qty, all institutions)", "Pairs"]
    _write_header(ws, row=4, headers=headers)
    for i, (isin, dev, n) in enumerate(rs, start=5):
        ws.cell(row=i, column=1, value=isin).font = DATA_FONT
        ws.cell(row=i, column=2, value=float(dev or 0)).font = DATA_FONT
        ws.cell(row=i, column=2).number_format = "#,##0.000000"
        ws.cell(row=i, column=3, value=n).font = DATA_FONT

    _autosize(ws, max_col=3)


# -------------------------------------------------- pair detail sheets

def _write_pairs_sheet(ws, repo: Repository, start: dt.date, end: dt.date,
                       *, status: str) -> None:
    rs = repo.db.execute("""
        SELECT
            m.match_id, m.partial_reason,
            i.source, i.transaction_type, i.isin, i.institution,
            i.quantity AS i_qty, e.quantity AS e_qty,
            i.trade_date AS i_trade, e.trade_date AS e_trade,
            i.settlement_date AS i_sett,
            e.source AS e_source,
            ABS(COALESCE(i.quantity,0) - COALESCE(e.quantity,0)) AS deviation_qty,
            m.matched_at, m.matched_by, m.reason
        FROM matches m
        JOIN transactions_all i ON m.internal_id = i.transaction_id
        LEFT JOIN transactions_all e ON m.external_id = e.transaction_id
        WHERE m.status = ?
          AND i.settlement_date BETWEEN ? AND ?
        ORDER BY i.isin, i.settlement_date
    """, [status, start, end]).fetchall()

    headers = [
        "Match ID", "Partial reason", "Internal source", "Type", "ISIN",
        "Institution", "Internal qty", "External qty",
        "Internal trade", "External trade", "Internal settlement",
        "External source", "Deviation (qty)", "Matched at", "Matched by", "Reason",
    ]
    _write_header(ws, row=1, headers=headers)
    for i, row in enumerate(rs, start=2):
        for col_idx, val in enumerate(row, start=1):
            c = ws.cell(row=i, column=col_idx,
                        value=_excel_safe(val))
            c.font = DATA_FONT

    _autosize(ws, max_col=len(headers))


# ------------------------------------------------------- unmatched sheet

def _write_unmatched_sheet(ws, repo: Repository, start: dt.date, end: dt.date) -> None:
    rs = repo.db.execute("""
        SELECT
            transaction_id, source, side, transaction_type, isin, institution,
            quantity, trade_date, settlement_date, native_reference, ingested_at
        FROM unmatched_pool
        WHERE settlement_date BETWEEN ? AND ?
        ORDER BY settlement_date, isin
    """, [start, end]).fetchall()

    headers = [
        "Transaction ID", "Source", "Side", "Type", "ISIN", "Institution",
        "Quantity", "Trade date", "Settlement date", "Native reference",
        "Ingested at",
    ]
    _write_header(ws, row=1, headers=headers)
    for i, row in enumerate(rs, start=2):
        for col_idx, val in enumerate(row, start=1):
            c = ws.cell(row=i, column=col_idx, value=_excel_safe(val))
            c.font = DATA_FONT
    _autosize(ws, max_col=len(headers))


# -------------------------------------------------------------- helpers

def _write_header(ws, *, row: int, headers: list[str]) -> None:
    for col_idx, h in enumerate(headers, start=1):
        c = ws.cell(row=row, column=col_idx, value=h)
        c.fill = HEADER_FILL
        c.font = HEADER_FONT
        c.alignment = Alignment(horizontal="left", vertical="center")


def _autosize(ws, *, max_col: int) -> None:
    for col_idx in range(1, max_col + 1):
        col_letter = get_column_letter(col_idx)
        max_len = 8
        for cell in ws[col_letter]:
            if cell.value is not None:
                max_len = max(max_len, min(len(str(cell.value)), 60))
        ws.column_dimensions[col_letter].width = max_len + 2


def _excel_safe(v):
    """openpyxl chokes on Decimal — coerce to float for numeric values."""
    if isinstance(v, Decimal):
        return float(v)
    if isinstance(v, dt.datetime):
        return v.replace(tzinfo=None)
    return v
